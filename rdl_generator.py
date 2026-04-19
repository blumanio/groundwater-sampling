"""
rdl_generator.py
Place in: server/rdl_generator.py
Usage: python3 rdl_generator.py '<json_string>'  → writes to stdout as base64
"""
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import io, os, sys, json, base64

LM = {0:None, 1:'thin', 2:'thin', 7:'hair'}
def bd(s): return Side(border_style=s) if s else Side()
def B(l=0,r=0,t=0,b=0):
    return Border(left=bd(LM.get(l)),right=bd(LM.get(r)),top=bd(LM.get(t)),bottom=bd(LM.get(b)))

def sc(ws,r,c,v,bold=False,sz=10,h='general',wrap=False,bl=0,br=0,bt=0,bb=0):
    cell=ws.cell(r,c)
    try: cell.value=v
    except: pass
    try:
        cell.font=Font(name='Arial',bold=bold,size=sz)
        cell.alignment=Alignment(horizontal=h,vertical='center',wrap_text=wrap)
        cell.border=B(bl,br,bt,bb)
    except: pass

def M(ws,r1,r2,c1,c2,v,**kw):
    ws.merge_cells(start_row=r1,start_column=c1,end_row=r2,end_column=c2)
    sc(ws,r1,c1,v,**kw)

def generate_rdl(data, assets_dir=None):
    wb=Workbook(); ws=wb.active; ws.title='Rdl'
    ws.page_setup.orientation='landscape'; ws.page_setup.paperSize=9

    for i,w in enumerate([9.9,27.1,13.4,13.6,31.0,10.6,21.4,18.6,24.1,25.9]):
        ws.column_dimensions[get_column_letter(i+1)].width=w
    ws.row_dimensions[1].height=35.2; ws.row_dimensions[2].height=30.8
    for r in range(3,52): ws.row_dimensions[r].height=14.2

    comm=data.get('committente','ENI  REWIND')
    M(ws,1,2,1,2,'',bl=0,br=2,bt=2,bb=2)
    M(ws,1,2,3,4,f'COMMITTENTE:\n{comm}',sz=8,h='center',wrap=True,bl=2,br=2,bt=2,bb=2)
    sc(ws,1,5,'Contratto:',bold=True,sz=8,h='center',wrap=True,bl=2,br=2,bt=2,bb=2)
    M(ws,1,1,6,7,data.get('data_str',''),bold=True,sz=11,h='center',bl=2,br=2,bt=2,bb=2)
    M(ws,1,2,8,9,f"{data.get('cantiere','')}\n{data.get('commessa','')}",bold=True,sz=8,h='left',wrap=True,bl=2,br=2,bt=2,bb=2)
    sc(ws,1,10,'RAPPORTO GIORNALIERO DEI LAVORI',sz=8,h='center',bl=2,br=2,bt=2,bb=2)
    sc(ws,2,5,f"Ordine Di Lavoro n.:\n{data.get('ordine_lavoro','')}",bold=True,sz=8,h='center',wrap=True,bl=2,br=2,bt=2,bb=2)
    M(ws,2,2,6,7,f"condizioni metereologiche: {data.get('meteo','sereno')}",bold=True,sz=8,h='center',wrap=True,bl=2,br=2,bt=2,bb=2)
    sc(ws,2,10,f"n. {data.get('n_rdl','1')}",bold=True,sz=11,h='right',bl=2,br=2,bt=2,bb=2)

    sc(ws,3,1,'DESCRIZIONE LAVORI',bold=True,sz=9,h='left',bl=2,br=0,bt=2,bb=2)
    for c in range(2,10): sc(ws,3,c,'',bl=0,br=0,bt=2,bb=2)
    sc(ws,3,10,'',bl=2,br=2,bt=2,bb=2)

    desc=data.get('descrizione_lavori',[])
    for i in range(13):
        r=4+i; txt=desc[i] if i<len(desc) else ''
        blv=2 if i%3!=1 else 0; btv=2 if i==0 else 0
        M(ws,r,r,1,9,txt,sz=11,h='left',bl=blv,br=1,bt=btv,bb=7)
        sc(ws,r,10,'',bl=1,br=2,bt=7,bb=7)

    M(ws,17,17,1,5,'FORNITURA   MATERIALI   IN   OPERA',bold=True,sz=9,h='center',bl=2,br=2,bt=2,bb=2)
    M(ws,17,17,6,10,'NOLEGGIO  MEZZI  E  ATTREZZATURE  DA  CANTIERE',bold=True,sz=9,h='center',bl=2,br=2,bt=2,bb=2)
    for c,l in [(1,'art.'),(2,'descrizione'),(3,'unità di mis.'),(4,'quantità'),(5,'impiego')]:
        sc(ws,18,c,l,sz=8,h='center',bl=2 if c==1 else 1,br=2 if c==5 else 1,bt=2,bb=1)
    sc(ws,18,6,'art.',sz=8,h='center',bl=0,br=1,bt=2,bb=1)
    M(ws,18,18,7,8,'descrizione',sz=8,h='center',bl=1,br=1,bt=2,bb=1)
    sc(ws,18,9,'ore',sz=8,h='center',bl=1,br=1,bt=2,bb=1)
    sc(ws,18,10,'impiego',sz=8,h='center',bl=1,br=2,bt=2,bb=1)

    noleggio=data.get('noleggio',[])
    mat_opera=data.get('materiali_opera',[])
    for i in range(7):
        r=19+i
        m=mat_opera[i] if i<len(mat_opera) else {}
        sc(ws,r,1,m.get('art',''),sz=10,bl=2,br=1,bt=7,bb=7)
        sc(ws,r,2,m.get('descrizione',''),sz=10,bl=1,br=1,bt=7,bb=7)
        sc(ws,r,3,m.get('unita',''),sz=10,bl=1,br=1,bt=7,bb=7)
        sc(ws,r,4,m.get('quantita',''),sz=10,h='center',bl=1,br=1,bt=7,bb=7)
        sc(ws,r,5,m.get('impiego',''),sz=10,bl=1,br=2,bt=7,bb=7)
        n=noleggio[i] if i<len(noleggio) else {}
        sc(ws,r,6,'',sz=10,bl=0,br=1,bt=7,bb=7)
        M(ws,r,r,7,8,n.get('descrizione',''),sz=10,h='left',bl=1,br=1,bt=7,bb=7)
        sc(ws,r,9,str(n.get('ore','')),sz=10,h='center',bl=1,br=1,bt=7,bb=7)
        sc(ws,r,10,n.get('impiego',''),sz=10,bl=1,br=2,bt=7,bb=7)

    M(ws,26,26,1,5,"FORNITURA   MATERIALI   A  PIE'  OPERA",bold=True,sz=9,h='center',bl=2,br=2,bt=2,bb=2)
    for c in range(6,11): sc(ws,26,c,'',bl=0 if c==6 else 1,br=2 if c==10 else 1,bt=7,bb=7)
    for c,l in [(1,'art.'),(2,'descrizione'),(3,'unità di mis.'),(4,'quantità'),(5,'impiego')]:
        sc(ws,27,c,l,sz=8,h='center',bl=2 if c==1 else 1,br=2 if c==5 else 1,bt=2,bb=1)
    for c in range(6,11): sc(ws,27,c,'',bl=0 if c==6 else 1,br=2 if c==10 else 1,bt=7,bb=7)

    mat_piedi=data.get('materiali_piedi',[])
    for i in range(6):
        r=28+i; m=mat_piedi[i] if i<len(mat_piedi) else {}
        sc(ws,r,1,m.get('art',''),sz=10,bl=2,br=1,bt=7,bb=7)
        sc(ws,r,2,m.get('descrizione',''),sz=10,bl=1,br=1,bt=7,bb=7)
        sc(ws,r,3,m.get('unita',''),sz=10,bl=1,br=1,bt=7,bb=7)
        sc(ws,r,4,m.get('quantita',''),sz=10,h='center',bl=1,br=1,bt=7,bb=7)
        sc(ws,r,5,m.get('impiego',''),sz=10,bl=1,br=2,bt=7,bb=7)

    M(ws,29,29,6,10,'NOTE LAVORI',bold=True,sz=8,h='left',bl=2,br=2,bt=2,bb=2)
    M(ws,28,28,6,10,'',bl=2,br=2,bt=0,bb=7)
    M(ws,30,33,6,10,data.get('note_lavori',''),sz=10,h='left',wrap=True,bl=2,br=2,bt=0,bb=2)

    M(ws,34,34,1,5,'LAVORI A MISURA',bold=True,sz=9,h='center',bl=2,br=2,bt=2,bb=2)
    M(ws,34,36,6,10,'',bl=2,br=2,bt=7,bb=7)
    for c,l in [(1,'art.'),(2,'descrizione'),(3,'unità di mis.'),(4,'quantità'),(5,'impiego')]:
        sc(ws,35,c,l,sz=8,h='center',bl=2 if c==1 else 1,br=2 if c==5 else 1,bt=2,bb=1)

    lav_mis=data.get('lavori_misura',[])
    for i in range(3):
        r=36+i; m=lav_mis[i] if i<len(lav_mis) else {}
        sc(ws,r,1,m.get('art',''),sz=8,bl=2,br=1,bt=7,bb=7)
        sc(ws,r,2,m.get('descrizione',''),sz=8,bl=1,br=1,bt=7,bb=7)
        sc(ws,r,3,m.get('unita',''),sz=8,bl=1,br=1,bt=7,bb=7)
        sc(ws,r,4,m.get('quantita',''),sz=8,h='center',bl=1,br=1,bt=7,bb=7)
        sc(ws,r,5,m.get('impiego',''),sz=10,bl=1,br=2,bt=7,bb=7)

    M(ws,39,39,6,10,'NOTE SICUREZZA',bold=True,sz=8,h='left',bl=2,br=2,bt=2,bb=2)
    for r in [37,38]: M(ws,r,r,6,10,'',bl=2,br=2,bt=7,bb=7)
    M(ws,40,41,6,10,data.get('note_sicurezza',''),sz=10,h='left',wrap=True,bl=2,br=2,bt=7,bb=7)

    M(ws,42,42,1,5,'PRESTAZIONI DI MANODOPERA IN ECONOMIA',bold=True,sz=9,h='center',bl=2,br=2,bt=2,bb=2)
    M(ws,42,42,6,10,'',bl=2,br=2,bt=7,bb=7)
    M(ws,43,43,1,2,'nominativo',sz=8,h='center',bl=2,br=1,bt=2,bb=1)
    sc(ws,43,3,'qualifica',sz=8,h='center',bl=0,br=1,bt=2,bb=1)
    sc(ws,43,4,'ore',sz=8,h='center',bl=1,br=1,bt=2,bb=1)
    sc(ws,43,5,'Impiego',sz=8,h='center',bl=1,br=2,bt=2,bb=1)
    M(ws,43,43,6,10,'',bl=2,br=2,bt=7,bb=7)

    manodopera=data.get('manodopera',[])
    for i in range(3):
        r=44+i; m=manodopera[i] if i<len(manodopera) else {}
        M(ws,r,r,1,2,m.get('nominativo',''),sz=9,h='center',bl=2,br=0,bt=1 if i==0 else 7,bb=7)
        sc(ws,r,3,m.get('qualifica',''),sz=8,h='center',bl=0,br=1,bt=7,bb=7)
        sc(ws,r,4,m.get('ore',''),sz=11,h='center',bl=1,br=1,bt=7,bb=7)
        sc(ws,r,5,m.get('impiego',''),sz=10,h='left',bl=1,br=2,bt=0,bb=7)
        M(ws,r,r,6,10,'',bl=2,br=2,bt=7,bb=2 if r==46 else 7)

    M(ws,47,50,6,7,' Committente',sz=10,h='center',bl=2,br=0,bt=2,bb=2)
    M(ws,47,47,8,10,'Contrattista',sz=10,h='left',bl=1,br=1,bt=2,bb=0)
    for r in [48,49,50]: M(ws,r,r,8,10,'',bold=True,sz=8,h='center',bl=1,br=1,bt=0,bb=2)
    for r in range(47,51):
        M(ws,r,r,1,2,'',sz=9,h='center',bl=2,br=1,bt=7,bb=7 if r<50 else 2)
        sc(ws,r,3,'',sz=8,bl=0,br=1,bt=7,bb=7 if r<50 else 2)
        sc(ws,r,4,'',sz=11,bl=1,br=1,bt=7,bb=7 if r<50 else 2)
        sc(ws,r,5,'',sz=10,bl=1,br=2,bt=0,bb=7 if r<50 else 2)

    # Images
    if assets_dir:
        for fname, anchor, w, h in [
            ('logo_eni.png','A1',55,50),
            ('logo_acr_strip.jpg','A2',170,35)
        ]:
            p = os.path.join(assets_dir, fname)
            if os.path.exists(p):
                try:
                    img=XLImage(p); img.width=w; img.height=h; img.anchor=anchor
                    ws.add_image(img)
                except: pass

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()

# SOSTITUISCI le ultime righe (if __name__=='__main__':) con:
if __name__=='__main__':
    import sys, json, base64
    # Read from stdin instead of argv (avoids Windows escaping issues)
    raw = sys.stdin.read()
    data = json.loads(raw) if raw.strip() else {}
    assets_dir = sys.argv[1] if len(sys.argv) > 1 else os.path.dirname(os.path.abspath(__file__))
    result = generate_rdl(data, assets_dir)
    sys.stdout.buffer.write(base64.b64encode(result))